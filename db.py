import telebot
from telebot import types
import webbrowser
import sqlite3
import openpyxl as op


rasp = 'rasp.xlsx'
raspy = 'raspy.xlsx'

rasp2 = op.load_workbook(rasp, data_only=True)
raspy2 = op.load_workbook(raspy, data_only=True)

sheet1 = rasp2.active
sheet2 = raspy2.active

max_rows_rasp = sheet1.max_row
max_rows_raspy = sheet2.max_row

abo = 2
kcla = 17
proj = 0




def rasp(message):
    markup_inline = types.InlineKeyboardMarkup()
    kbtn1 = types.InlineKeyboardButton('5а', callback_data='5а')
    kbtn2 = types.InlineKeyboardButton('5б', callback_data='5б')
    kbtn3 = types.InlineKeyboardButton('5в', callback_data='5в')
    kbtn4 = types.InlineKeyboardButton('5г', callback_data='5г')
    kbtn5 = types.InlineKeyboardButton('5д', callback_data='5д')
    kbtn6 = types.InlineKeyboardButton('5е', callback_data='5е')
    kbtn7 = types.InlineKeyboardButton('5ж', callback_data='5ж')
    kbtn8 = types.InlineKeyboardButton('5з', callback_data='5з')
    kbtn9 = types.InlineKeyboardButton('6а', callback_data='6a')
    kbtn10 = types.InlineKeyboardButton('6б', callback_data='6б')
    kbtn11 = types.InlineKeyboardButton('6в', callback_data='6в')
    kbtn12 = types.InlineKeyboardButton('6г', callback_data='6г')
    kbtn13 = types.InlineKeyboardButton('6д', callback_data='6д')
    kbtn14 = types.InlineKeyboardButton('6е', callback_data='6е')
    kbtn15 = types.InlineKeyboardButton('7а', callback_data='7а')
    kbtn16 = types.InlineKeyboardButton('7б', callback_data='7б')
    kbtn17 = types.InlineKeyboardButton('7в', callback_data='7в')
    kbtn18 = types.InlineKeyboardButton('7г', callback_data='7г')
    kbtn19 = types.InlineKeyboardButton('7д', callback_data='7д')
    kbtn20 = types.InlineKeyboardButton('7е', callback_data='7е')
    kbtn21 = types.InlineKeyboardButton('7ж', callback_data='7ж')
    kbtn22 = types.InlineKeyboardButton('7з', callback_data='7з')
    kbtn23 = types.InlineKeyboardButton('7и', callback_data='7и')
    kbtn24 = types.InlineKeyboardButton('7к', callback_data='7к')
    kbtn25 = types.InlineKeyboardButton('8а', callback_data='8а')
    kbtn26 = types.InlineKeyboardButton('8б', callback_data='8б')
    kbtn27 = types.InlineKeyboardButton('8в', callback_data='8в')
    kbtn28 = types.InlineKeyboardButton('8г', callback_data='8г')
    kbtn29 = types.InlineKeyboardButton('8д', callback_data='8д')
    kbtn30 = types.InlineKeyboardButton('8е', callback_data='8е')
    kbtn31 = types.InlineKeyboardButton('9а', callback_data='9а')
    kbtn32 = types.InlineKeyboardButton('9б', callback_data='9б')
    kbtn33 = types.InlineKeyboardButton('9в', callback_data='9в')
    kbtn34 = types.InlineKeyboardButton('9г', callback_data='9г')
    kbtn35 = types.InlineKeyboardButton('9д', callback_data='9д')
    kbtn36 = types.InlineKeyboardButton('9е', callback_data='9е')
    kbtn37 = types.InlineKeyboardButton('10а', callback_data='10а')
    kbtn38 = types.InlineKeyboardButton('10б', callback_data='10б')
    kbtn39 = types.InlineKeyboardButton('10в', callback_data='10в')
    kbtn40 = types.InlineKeyboardButton('11а', callback_data='11а')
    kbtn41 = types.InlineKeyboardButton('11б', callback_data='11б')
    kbtn42 = types.InlineKeyboardButton('5 класс', callback_data='5')
    kbtn43 = types.InlineKeyboardButton('6 класс', callback_data='6')
    kbtn44 = types.InlineKeyboardButton('7 класс', callback_data='7')
    kbtn45 = types.InlineKeyboardButton('8 класс', callback_data='8')
    kbtn46 = types.InlineKeyboardButton('9 класс', callback_data='9')
    kbtn47 = types.InlineKeyboardButton('10 класс', callback_data='10')
    kbtn48 = types.InlineKeyboardButton('11 класс', callback_data='11')
    markup_inline.row(kbtn42, kbtn43)
    markup_inline.row(kbtn44, kbtn45)
    markup_inline.row(kbtn46)
    markup_inline.row(kbtn47)
    markup_inline.row(kbtn48)
    bot.send_message(message.chat.id, 'Выберите класс:', reply_markup=markup_inline)




    for i in range(2, max_rows_rasp // 2 + 1):
        sku = sheet1.cell(row=i * 2, column=abo * kcla).value
        clas = sheet1.cell(row=3, column=abo * kcla).value
        global proj
        proj += 1

        if not sku:
            sku = 'нет урока'

        bot.send_message(message.chat.id, f'{proj}, {sku}')
    bot.send_message(message.chat.id, f'{clas}')
    proj = 0


@bot.message_handler()