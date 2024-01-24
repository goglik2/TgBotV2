import telebot
from telebot import types
import webbrowser
import sqlite3
import openpyxl as op
import os

bot = telebot.TeleBot('6741433926:AAFAOv4jNejzzQD_Gs7KxLtXA-xluG8jAcU')

users = []
c_users = 0
rasp = 'rasp.xlsx'

joinedFile = open('ids.txt', 'r')
joinedUsers = set()
for line in joinedFile:
    joinedUsers.add(line.strip())
joinedFile.close()

rasp2 = op.load_workbook(rasp, data_only=True)

sheet1 = rasp2.active

max_rows_rasp = sheet1.max_row

abo = 2
kcla = 1
proj = 0





@bot.message_handler(commands=['postrasp23'])
def postrasp(message):
    markup_inline = types.InlineKeyboardMarkup()
    rbtn1 = types.InlineKeyboardButton('Да', callback_data='RaspYes')
    rbtn2 = types.InlineKeyboardButton('Нет', callback_data='RaspNo')
    markup_inline.row(rbtn1, rbtn2)
    bot.send_message(message.chat.id, 'Вы хотите загрузить рассписание?', reply_markup=markup_inline)



@bot.message_handler(commands=['post23'])
def post(message):
    with open('ids.txt', 'r') as file:
        lines = file.readlines()

    unique_lines = set(lines)
    with open('ids.txt', 'w') as file:
        file.writelines(unique_lines)

    for user in joinedUsers:
        bot.send_message(user, message.text[message.text.find(' '):])

@bot.message_handler(commands=['mg'])    #команда только для разработчиков, нужна для просмотра информации о чате и пользователе
def mg(message):
    with open('ids.txt', 'r') as file:
        lines = file.readlines()

    unique_lines = set(lines)
    with open('ids.txt', 'w') as file:
        file.writelines(unique_lines)

    bot.send_message(message.chat.id, message)
    bot.send_message(message.chat.id, f'Все пользователи:')
    for item in users:
        bot.send_message(message.chat.id, f'{item}')

    bot.send_message(message.chat.id, f'Всего пользователей: {c_users}')
    joinedFile = open('ids.txt', 'r')
    bot.send_document(message.chat.id, joinedFile)
    joinedFile.close()

@bot.message_handler(commands=['start'])  #Команда для запуска бота и его стартовой функции
def start(message):
    if not str(message.chat.id) in joinedUsers:
        joinedFile = open('ids.txt', 'a')
        joinedFile.write(str(message.chat.id) + '\n')
        joinedUsers.add(message.chat.id)

    user_id = [message.from_user.first_name, message.from_user.last_name, message.from_user.username]            #Заполнение списка всех пользователей бота
    if user_id not in users:
        users.append(user_id)
        global c_users
        c_users += 1

    bot.send_message(message.chat.id, f'Приветсвую, {message.from_user.first_name}')
    markup = types.ReplyKeyboardMarkup()
    btn1 = types.KeyboardButton('Расписание')   #имена для кнопок
    btn2 = types.KeyboardButton('Помощь')        #имена для кнопок
    btn3 = types.KeyboardButton('Перезапустить')  #имена для кнопок
    markup.row(btn1)     #Создаёт кнопку под вводом 1 ряд
    markup.row(btn2)     #Создаёт кнопку под вводом 2 ряд
    markup.row(btn3)     #Создаёт кнопку под вводом 3 ряд
    bot.send_message(message.chat.id, f'Список команд для этого бота:\n/start - перезапустить\n/help - список команд\n/rasp - Расписание', reply_markup=markup)



@bot.message_handler(commands=['help'])        #при вводе команды help вылезает сообщение
def info(message):
    bot.send_message(message.chat.id,f'Список команд для этого бота:\n/start - перезапустить\n/help - список команд\n/rasp - Расписание')


@bot.message_handler(commands=['rasp'])        #при вводе команды rasp вылезает сообщение
def rasp(message):
    markup_inline = types.InlineKeyboardMarkup()
    kbtn42 = types.InlineKeyboardButton('5 класс', callback_data='5')      #даём имя кнопке, присваиваем ей класс и коллбэк дату
    kbtn43 = types.InlineKeyboardButton('6 класс', callback_data='6')
    kbtn44 = types.InlineKeyboardButton('7 класс', callback_data='7')
    kbtn45 = types.InlineKeyboardButton('8 класс', callback_data='8')
    kbtn46 = types.InlineKeyboardButton('9 класс', callback_data='9')
    kbtn47 = types.InlineKeyboardButton('10 класс', callback_data='10')
    kbtn48 = types.InlineKeyboardButton('11 класс', callback_data='11')
    markup_inline.row(kbtn42, kbtn43)             #собсна выводим то шо сделали
    markup_inline.row(kbtn44, kbtn45)
    markup_inline.row(kbtn46)
    markup_inline.row(kbtn47)
    markup_inline.row(kbtn48)
    bot.send_message(message.chat.id, 'Выберите класс:', reply_markup= markup_inline)

@bot.message_handler(func=lambda message: True)         #постоянная работа функции
def on_click(message):            #по сути дублирование всех команд но для работы нужна кнопка
    if message.text == 'Расписание':
        markup_inline = types.InlineKeyboardMarkup()
        kbtn42 = types.InlineKeyboardButton('5 класс', callback_data='5')        #имя,тип,текст и кол бэк дата
        kbtn43 = types.InlineKeyboardButton('6 класс', callback_data='6')
        kbtn44 = types.InlineKeyboardButton('7 класс', callback_data='7')
        kbtn45 = types.InlineKeyboardButton('8 класс', callback_data='8')
        kbtn46 = types.InlineKeyboardButton('9 класс', callback_data='9')
        kbtn47 = types.InlineKeyboardButton('10 класс', callback_data='10')
        kbtn48 = types.InlineKeyboardButton('11 класс', callback_data='11')
        markup_inline.row(kbtn42, kbtn43)
        markup_inline.row(kbtn44, kbtn45)
        markup_inline.row(kbtn46)
        markup_inline.row(kbtn47)
        markup_inline.row(kbtn48)
        bot.send_message(message.chat.id, 'Выберите класс:', reply_markup = markup_inline)     #воводятся с этим текстом

    elif message.text == 'Помощь':
        bot.send_message(message.chat.id, f'Список команд для этого бота:\n/start - перезапустить\n/help - список команд\n/rasp - Расписание')

    elif message.text == 'Разработчик':
        bot.send_message(message.chat.id, 'Это Гриша сделал(и немного вова)')

    elif message.text == 'Перезапустить':
        if not str(message.chat.id) in joinedUsers:
            joinedFile = open('ids.txt', 'a')
            joinedFile.write(str(message.chat.id) + '\n')
            joinedUsers.add(message.chat.id)

        user_id = [message.from_user.first_name, message.from_user.last_name,
                   message.from_user.username]  # Заполнение списка всех пользователей бота
        if user_id not in users:
            users.append(user_id)
            global c_users
            c_users += 1

        bot.send_message(message.chat.id, f'Приветсвую, {message.from_user.first_name}')
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Расписание')  # имена для кнопок
        btn2 = types.KeyboardButton('Помощь')  # имена для кнопок
        btn3 = types.KeyboardButton('Перезапустить')  # имена для кнопок
        markup.row(btn1)  # Создаёт кнопку под вводом 1 ряд
        markup.row(btn2)  # Создаёт кнопку под вводом 2 ряд
        markup.row(btn3)  # Создаёт кнопку под вводом 3 ряд
        bot.send_message(message.chat.id, f'Список команд для этого бота:\n/start - перезапустить\n/help - список команд\n/rasp - Расписание', reply_markup=markup)

@bot.callback_query_handler(func=lambda call: True)     #про это я писал
def clasrasp(call):       #тут идёт обращение к калу
    if call.data == '5':       # если это пришло то выполняется действия ниже и тд
        markup_inline = types.InlineKeyboardMarkup()
        kbtn1 = types.InlineKeyboardButton('5а', callback_data='5а')
        kbtn2 = types.InlineKeyboardButton('5б', callback_data='5б')
        kbtn3 = types.InlineKeyboardButton('5в', callback_data='5в')
        kbtn4 = types.InlineKeyboardButton('5г', callback_data='5г')
        kbtn5 = types.InlineKeyboardButton('5д', callback_data='5д')
        kbtn6 = types.InlineKeyboardButton('5е', callback_data='5е')
        kbtn7 = types.InlineKeyboardButton('5ж', callback_data='5ж')
        kbtn8 = types.InlineKeyboardButton('5з', callback_data='5з')
        markup_inline.row(kbtn1, kbtn2)
        markup_inline.row(kbtn3, kbtn4)
        markup_inline.row(kbtn5, kbtn6)
        markup_inline.row(kbtn7, kbtn8)
        bot.send_message(call.message.chat.id, f'Выберите букву', reply_markup = markup_inline)

    elif call.data == '6':
        markup_inline = types.InlineKeyboardMarkup()
        kbtn9 = types.InlineKeyboardButton('6а', callback_data='6a')
        kbtn10 = types.InlineKeyboardButton('6б', callback_data='6б')
        kbtn11 = types.InlineKeyboardButton('6в', callback_data='6в')
        kbtn12 = types.InlineKeyboardButton('6г', callback_data='6г')
        kbtn13 = types.InlineKeyboardButton('6д', callback_data='6д')
        kbtn14 = types.InlineKeyboardButton('6е', callback_data='6е')
        markup_inline.row(kbtn9, kbtn10)
        markup_inline.row(kbtn11, kbtn12)
        markup_inline.row(kbtn13, kbtn14)
        bot.send_message(call.message.chat.id, f'Выберите букву', reply_markup = markup_inline)

    elif call.data == '7':
        markup_inline = types.InlineKeyboardMarkup()
        kbtn15 = types.InlineKeyboardButton('7а', callback_data='7а')
        kbtn16 = types.InlineKeyboardButton('7б', callback_data='7б')
        kbtn17 = types.InlineKeyboardButton('7в', callback_data='7в')
        kbtn18 = types.InlineKeyboardButton('7г', callback_data='7г')
        kbtn19 = types.InlineKeyboardButton('7д', callback_data='7д')
        kbtn20 = types.InlineKeyboardButton('7е', callback_data='7е')
        kbtn21 = types.InlineKeyboardButton('7ж', callback_data='7ж')
        kbtn22 = types.InlineKeyboardButton('7з', callback_data='7з')
        kbtn23 = types.InlineKeyboardButton('7и', callback_data='7и')
        kbtn24 = types.InlineKeyboardButton('7к', callback_data='7к')
        markup_inline.row(kbtn15, kbtn16)
        markup_inline.row(kbtn17, kbtn18)
        markup_inline.row(kbtn19, kbtn20)
        markup_inline.row(kbtn21, kbtn22)
        markup_inline.row(kbtn23, kbtn24)
        bot.send_message(call.message.chat.id, f'Выберите букву', reply_markup = markup_inline)

    elif call.data == '8':
        markup_inline = types.InlineKeyboardMarkup()
        kbtn25 = types.InlineKeyboardButton('8а', callback_data='8а')
        kbtn26 = types.InlineKeyboardButton('8б', callback_data='8б')
        kbtn27 = types.InlineKeyboardButton('8в', callback_data='8в')
        kbtn28 = types.InlineKeyboardButton('8г', callback_data='8г')
        kbtn29 = types.InlineKeyboardButton('8д', callback_data='8д')
        kbtn30 = types.InlineKeyboardButton('8е', callback_data='8е')
        markup_inline.row(kbtn25, kbtn26)
        markup_inline.row(kbtn27, kbtn28)
        markup_inline.row(kbtn29, kbtn30)
        bot.send_message(call.message.chat.id, f'Выберите букву', reply_markup = markup_inline)

    elif call.data == '9':
        markup_inline = types.InlineKeyboardMarkup()
        kbtn31 = types.InlineKeyboardButton('9а', callback_data='9а')
        kbtn32 = types.InlineKeyboardButton('9б', callback_data='9б')
        kbtn33 = types.InlineKeyboardButton('9в', callback_data='9в')
        kbtn34 = types.InlineKeyboardButton('9г', callback_data='9г')
        kbtn35 = types.InlineKeyboardButton('9д', callback_data='9д')
        kbtn36 = types.InlineKeyboardButton('9е', callback_data='9е')
        markup_inline.row(kbtn31, kbtn32)
        markup_inline.row(kbtn33, kbtn34)
        markup_inline.row(kbtn35, kbtn36)
        bot.send_message(call.message.chat.id, f'Выберите букву', reply_markup=markup_inline)

    elif call.data == '10':
        markup_inline = types.InlineKeyboardMarkup()
        kbtn37 = types.InlineKeyboardButton('10а', callback_data='10а')
        kbtn38 = types.InlineKeyboardButton('10б', callback_data='10б')
        kbtn39 = types.InlineKeyboardButton('10в', callback_data='10в')
        markup_inline.row(kbtn37, kbtn38)
        markup_inline.row(kbtn39)
        bot.send_message(call.message.chat.id, f'Выберите букву', reply_markup=markup_inline)

    elif call.data == '11':
        markup_inline = types.InlineKeyboardMarkup()
        kbtn40 = types.InlineKeyboardButton('11а', callback_data='11а')
        kbtn41 = types.InlineKeyboardButton('11б', callback_data='11б')
        markup_inline.row(kbtn40, kbtn41)
        bot.send_message(call.message.chat.id, f'Выберите букву', reply_markup=markup_inline)

    elif call.data == 'RaspYes':
        bot.send_message(call.message.chat.id, 'Скоро добавлю!')

    elif call.data == 'RaspNo':
        bot.send_message(call.message.chat.id, 'Ну ладно')


    elif call.data == 'RaspPubl':
        with open('ids.txt', 'r') as file:
            lines = file.readlines()

        unique_lines = set(lines)
        with open('ids.txt', 'w') as file:
            file.writelines(unique_lines)

        for user in joinedUsers:
            bot.send_message(user, 'Рассписание обновилось!')



    #тута тестовый образец для вывода рассписание из файла
    # for i in range(2, max_rows_rasp // 2 + 1):
    # sku = sheet1.cell(row=i * 2, column=abo * kcla).value
    # clas = sheet1.cell(row=3, column=abo * kcla).value
    # global proj
    # proj += 1

    # if not sku:
    # sku = 'нет урока'

    # bot.send_message(message.chat.id, f'{proj}, {sku}')
    # bot.send_message(message.chat.id, f'{clas}')
    # proj = 0
    # rasp = open('./rasp.xlsx', 'rb')
    # bot.send_document(message.chat.id, rasp)

bot.polling(none_stop=True)